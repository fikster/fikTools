from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles import PatternFill


def cell_name(row, column):
    return '{}{}'.format(get_column_letter(column), row)


def cell_range(first_row, last_row, first_column, last_column):
    return '{}:{}'.format(cell_name(first_row, first_column), cell_name(last_row, last_column))


def title(ws, row, headers, width=[]):
    column = 1
    for i in range(0, len(headers)):
        set_cell(ws, row, column, headers[i], is_bold=True, horizontal_alignment='center', vertical_alignment='center')
        if i < len(width) and width[i] != -1:
            set_column_width(ws, column, width[i])
        column += 1


def set_cell(ws, row, column, value, span=1, size=14, is_bold=False, is_red=False, horizontal_alignment='left', vertical_alignment='top', wrap_text=True):
    if horizontal_alignment in ['middle', 'centre']:
        horizontal_alignment = 'center'

    if vertical_alignment in ['middle', 'centre']:
        vertical_alignment = 'center'

    ws.cell(row=row, column=column).value = value
    ws.cell(row=row, column=column).alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment, wrap_text=wrap_text)

    if is_bold:

        if is_red:
            ws.cell(row=row, column=column).font = Font(name='Gill Sans MT', size=size, bold=True, color='FF0000')
        else:
            ws.cell(row=row, column=column).font = Font(name='Gill Sans MT', size=size, bold=True)

    else:

        if is_red:
            ws.cell(row=row, column=column).font = Font(name='Gill Sans MT', size=size, color='FF0000')
        else:
            ws.cell(row=row, column=column).font = Font(name='Gill Sans MT', size=size)

    if span != 1:
        ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + span -1)


def get_cell_value(ws, row, column):
    return ws.cell(row=row, column=column).value if ws.cell(row=row, column=column).value is not None else ''


def set_cell_background_colour(ws, row, column, colour):
    ws.cell(row=row, column=column).fill = PatternFill(start_color=colour, end_color=colour, fill_type='solid')


def set_cell_background_user_entry(ws, row, column):
    set_cell_background_colour(ws, row, column, colour='FFF3B2')


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def set_column_width(ws, column, width):
    ws.column_dimensions[get_column_letter(column)].width = width


def set_borders(ws, first_row, last_row, first_column, last_column):

    # Grid

    for r in range(first_row, last_row + 1):
        for c in range(first_column, last_column + 1):
            ws.cell(row=r, column=c).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
            )

    # Borders

    for r in range(first_row, last_row + 1):
        if first_column == last_column:
            ws.cell(row=r, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='medium'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
            )
        else:
            ws.cell(row=r, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
            )
            ws.cell(row=r, column=last_column).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='medium'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
            )

    for c in range(first_column, last_column + 1):
        if first_row == last_row:
            ws.cell(row=first_row, column=c).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='medium'),
            )
        else:
            ws.cell(row=first_row, column=c).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='thin'),
            )
            ws.cell(row=last_row, column=c).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='medium'),
            )

    # Corners

    if first_row == last_row:

        if first_column == last_column:
            ws.cell(row=first_row, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='medium'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='medium'),
            )
        else:
            ws.cell(row=first_row, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='thin'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='medium'),
            )
            ws.cell(row=first_row, column=last_column).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='medium'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='medium'),
            )
    else:
        if first_column == last_column:
            ws.cell(row=first_row, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='medium'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='thin'),
            )
            ws.cell(row=last_row, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='medium'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='medium'),
            )
        else:
            ws.cell(row=first_row, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='thin'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='thin'),
            )
            ws.cell(row=first_row, column=last_column).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='medium'),
                top=Side(border_style='medium'),
                bottom=Side(border_style='thin'),
            )
            ws.cell(row=last_row, column=first_column).border = Border(
                left=Side(border_style='medium'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='medium'),
            )
            ws.cell(row=last_row, column=last_column).border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='medium'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='medium'),
            )


def delete_extra_sheet(wb, sheets_to_keep):

    for ws in wb:
        if ws.title.lower() not in sheets_to_keep:
            wb.remove_sheet(ws)
